#!/usr/bin/env python3
"""
USCP 경쟁 입찰 견적서 생성 스크립트
- 부팅 주식회사: 부가세 포함 63,000,000원
- 앤오픈: 부가세 포함 67,000,000원

(우리 회사 투지넥스트: 부가세 포함 55,000,000원 — 최저가 낙찰)
"""
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter


def thin_border():
    s = Side(style='thin', color='999999')
    return Border(left=s, right=s, top=s, bottom=s)


def create_quote(output_path, company, quote_date, kor_amount, modules,
                  maintenance, supply_total, vat, grand_total):
    wb = Workbook()
    wb.remove(wb.active)

    # === 산출내역 시트 ===
    ws = wb.create_sheet("산출내역")

    big = Font(name='맑은 고딕', size=18, bold=True)
    h1 = Font(name='맑은 고딕', size=11, bold=True)
    norm = Font(name='맑은 고딕', size=10)
    bold = Font(name='맑은 고딕', size=10, bold=True)
    big_bold = Font(name='맑은 고딕', size=12, bold=True, color='C00000')

    center = Alignment(horizontal='center', vertical='center', wrap_text=True)
    right = Alignment(horizontal='right', vertical='center')
    left = Alignment(horizontal='left', vertical='center', wrap_text=True)

    header_fill = PatternFill('solid', fgColor='D9E1F2')
    sub_fill = PatternFill('solid', fgColor='FFF2CC')
    total_fill = PatternFill('solid', fgColor='FFD966')
    grand_fill = PatternFill('solid', fgColor='F4B084')

    border = thin_border()

    # Title
    r = 1
    ws.merge_cells(f'A{r}:H{r}')
    c = ws.cell(r, 1)
    c.value = '견   적   서'
    c.font = big
    c.alignment = center
    ws.row_dimensions[r].height = 36
    r += 2

    # Header lines (left + right info)
    header_rows = [
        (f"견적일자 : {quote_date}", f" 사업자등록번호 : {company['biz_no']}", None),
        ("수      신 : 공주대학교 지역사회특화센터 한혜진님",
         f" 상   호 : {company['name']}", f" 대   표 : {company['ceo']}"),
        ("건      명 : USCP 플랫폼 구축 및 3년차 유지보수비 산정",
         f" 소재지 : {company['address']}", None),
        (f"담  당 자 : {company['contact']}",
         f" 업   태 : {company['business']}", f" 종  목 : {company['type']}"),
        ("견적유효일: 견적일로 부터 1개월",
         f" 전   화 : {company['phone']}", f" 팩  스 : {company.get('fax', '')}"),
    ]

    for left_text, right_text, right2_text in header_rows:
        ws.cell(r, 1).value = left_text
        ws.cell(r, 1).font = norm
        ws.cell(r, 1).alignment = left
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=4)

        ws.cell(r, 5).value = right_text
        ws.cell(r, 5).font = norm
        ws.cell(r, 5).alignment = left
        ws.merge_cells(start_row=r, start_column=5, end_row=r, end_column=6)

        if right2_text:
            ws.cell(r, 7).value = right2_text
            ws.cell(r, 7).font = norm
            ws.cell(r, 7).alignment = left
            ws.merge_cells(start_row=r, start_column=7, end_row=r, end_column=8)
        r += 1
    r += 1

    # Grand total banner
    ws.merge_cells(f'A{r}:H{r}')
    c = ws.cell(r, 1)
    c.value = f"일금 : {kor_amount} (₩{grand_total:,}_VAT 포함)"
    c.font = big_bold
    c.alignment = center
    c.fill = grand_fill
    ws.row_dimensions[r].height = 28
    r += 2

    # Table header
    hdr = ['번호', '구분', '내   역', '비   고', 'M/M', '단가', '공급금액', '비고']
    for i, h in enumerate(hdr, 1):
        c = ws.cell(r, i)
        c.value = h
        c.font = h1
        c.alignment = center
        c.fill = header_fill
        c.border = border
    ws.row_dimensions[r].height = 24
    r += 1

    # Sub-header row (project)
    ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=8)
    c = ws.cell(r, 1)
    c.value = '#'
    c.font = bold
    c.alignment = center
    c.border = border
    c2 = ws.cell(r, 2)
    c2.value = 'USCP 플랫폼 구축 및 3년차 유지보수비 산정'
    c2.font = bold
    c2.alignment = left
    c2.border = border
    for i in range(3, 9):
        ws.cell(r, i).border = border
    r += 1

    # Build items
    num = 1

    def write_item(num, item, fill=None):
        nonlocal r
        cells_vals = [
            num,
            item.get('category', ''),
            item.get('name', ''),
            item.get('detail', ''),
            item.get('mm', ''),
            item.get('unit_price', ''),
            item.get('amount', ''),
            item.get('note', ''),
        ]
        for i, v in enumerate(cells_vals, 1):
            c = ws.cell(r, i)
            c.value = v
            c.font = norm
            c.border = border
            if i in (1, 5):
                c.alignment = center
            elif i in (6, 7):
                c.alignment = right
            else:
                c.alignment = left
            if fill:
                c.fill = fill
        # Numeric formatting
        if isinstance(item.get('unit_price'), (int, float)):
            ws.cell(r, 6).number_format = '#,##0'
        if isinstance(item.get('amount'), (int, float)):
            ws.cell(r, 7).number_format = '#,##0'
        # Row height for wrapping
        ws.row_dimensions[r].height = max(36,
            18 * max(item.get('name', '').count('\n') + 1,
                     item.get('detail', '').count('\n') + 1))
        r += 1

    def write_subtotal(label, detail, mm, amount):
        nonlocal r
        ws.cell(r, 1).value = '-'
        ws.cell(r, 2).value = '소계'
        ws.cell(r, 3).value = label
        ws.cell(r, 4).value = detail
        ws.cell(r, 5).value = mm
        ws.cell(r, 7).value = amount
        for i in range(1, 9):
            ws.cell(r, i).font = bold
            ws.cell(r, i).border = border
            ws.cell(r, i).fill = sub_fill
        ws.cell(r, 1).alignment = center
        ws.cell(r, 2).alignment = center
        ws.cell(r, 3).alignment = left
        ws.cell(r, 4).alignment = left
        ws.cell(r, 5).alignment = center
        ws.cell(r, 7).alignment = right
        ws.cell(r, 7).number_format = '#,##0'
        r += 1

    # Modules
    for item in modules:
        write_item(num, item)
        num += 1

    write_subtotal('구축비 소계', '분석·설계 + 개발 모듈 + 시범운영',
                   sum(m['mm'] for m in modules),
                   sum(m['amount'] for m in modules))

    # Maintenance
    for item in maintenance:
        write_item(num, item)
        num += 1

    write_subtotal('유지보수 소계 (2~3년차)', '',
                   sum(m['mm'] for m in maintenance),
                   sum(m['amount'] for m in maintenance))

    # 공급가 합계
    ws.cell(r, 1).value = '-'
    ws.cell(r, 2).value = '합계'
    ws.cell(r, 3).value = '3년 통합 공급가 합계'
    ws.cell(r, 4).value = '구축 + 유지보수 1~3년차'
    ws.cell(r, 5).value = sum(m['mm'] for m in modules) + sum(m['mm'] for m in maintenance)
    ws.cell(r, 7).value = supply_total
    for i in range(1, 9):
        ws.cell(r, i).font = bold
        ws.cell(r, i).border = border
        ws.cell(r, i).fill = total_fill
    ws.cell(r, 1).alignment = center
    ws.cell(r, 2).alignment = center
    ws.cell(r, 3).alignment = left
    ws.cell(r, 4).alignment = left
    ws.cell(r, 5).alignment = center
    ws.cell(r, 7).alignment = right
    ws.cell(r, 7).number_format = '#,##0'
    r += 1

    # VAT
    ws.cell(r, 1).value = '-'
    ws.cell(r, 3).value = 'VAT 10% (별도)'
    ws.cell(r, 7).value = vat
    for i in range(1, 9):
        ws.cell(r, i).font = bold
        ws.cell(r, i).border = border
        ws.cell(r, i).fill = total_fill
    ws.cell(r, 1).alignment = center
    ws.cell(r, 3).alignment = left
    ws.cell(r, 7).alignment = right
    ws.cell(r, 7).number_format = '#,##0'
    r += 1

    # 부가세 포함 총액
    ws.cell(r, 1).value = ''
    ws.cell(r, 2).value = '합계'
    ws.cell(r, 3).value = '부가세 포함 총액'
    ws.cell(r, 7).value = grand_total
    for i in range(1, 9):
        ws.cell(r, i).font = big_bold
        ws.cell(r, i).border = border
        ws.cell(r, i).fill = grand_fill
    ws.cell(r, 2).alignment = center
    ws.cell(r, 3).alignment = left
    ws.cell(r, 7).alignment = right
    ws.cell(r, 7).number_format = '#,##0'
    ws.row_dimensions[r].height = 26
    r += 2

    # Footer
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=8)
    ws.cell(r, 1).value = '끝.'
    ws.cell(r, 1).alignment = Alignment(horizontal='right', vertical='center')
    ws.cell(r, 1).font = bold

    # Column widths
    widths = [6, 10, 32, 38, 6, 14, 16, 18]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    wb.save(output_path)
    print(f"✓ Created: {output_path}")


# ────────────────────────────────────────────
# 부팅 주식회사 (6,300만 원, VAT 포함)
# ────────────────────────────────────────────
booting_company = {
    'name': '부팅 주식회사',
    'biz_no': '234-86-12345',
    'ceo': '한 정 민',
    'address': '서울특별시 서초구 강남대로 123, 5층',
    'business': '정보통신업',
    'type': '시스템 통합 자문 및 구축',
    'phone': '02-555-1234',
    'fax': '02-555-1235',
    'contact': '김부장 (010-2345-6789)',
}

booting_modules = [
    {'category': '분석·설계',
     'name': '사전 준비 및 분석',
     'detail': '요구사항 분석\nUI 및 메뉴 설계\nDB 설계\n소개 페이지 작성\nUX 워크숍·시장 조사',
     'mm': 1.6, 'unit_price': 4500000, 'amount': 7200000,
     'note': 'PM·아키텍트·UI/UX 협업 (고급)\n+ UX 조사 보강'},
    {'category': '개발',
     'name': '회원·인증 모듈',
     'detail': '회원가입\n로그인\n비밀번호·프로필 변경\n회원 탈퇴\n이용약관·개인정보처리방침',
     'mm': 0.7, 'unit_price': 3800000, 'amount': 2660000,
     'note': '풀스택 중급'},
    {'category': '개발',
     'name': '지역문제 제보·게이트키핑 모듈',
     'detail': '제보·공감투표·댓글\n6단계 워크플로우\n다단계 일괄 처리\n이메일 알림',
     'mm': 2.0, 'unit_price': 3800000, 'amount': 7600000,
     'note': '핵심 워크플로우\n(일괄 처리 강화)'},
    {'category': '개발',
     'name': '리빙랩 운영 + 프로젝트 게시판',
     'detail': '등록·활동 타임라인\n산출물 DB\n성공사례 4단계\n멤버 전용 게시판',
     'mm': 1.5, 'unit_price': 3800000, 'amount': 5700000,
     'note': '게시판 포함'},
    {'category': '개발',
     'name': '멘토단·학생팀 매칭 모듈',
     'detail': '멘토 자격 부여\n학생팀 구성\n팀원 매칭 관리\n이메일 통보',
     'mm': 0.8, 'unit_price': 3800000, 'amount': 3040000,
     'note': '풀스택 중급'},
    {'category': '개발',
     'name': '협력 네트워크 모듈',
     'detail': '협력기관\nMOU\n프로그램\n커뮤니티',
     'mm': 0.5, 'unit_price': 3800000, 'amount': 1900000,
     'note': '풀스택 중급'},
    {'category': '개발',
     'name': '성과자료 모듈',
     'detail': '성과지표 연동\n실적 입력·자동 집계\n엑셀 다운로드',
     'mm': 0.6, 'unit_price': 3800000, 'amount': 2280000,
     'note': '기본'},
    {'category': '개발',
     'name': '콘텐츠 관리 모듈 + 이벤트 캘린더',
     'detail': '게시판 생성(공지·이벤트·자료실·배너·약관 버전)\n이벤트 캘린더\nWYSIWYG 에디터',
     'mm': 1.3, 'unit_price': 3800000, 'amount': 4940000,
     'note': '추가 옵션 포함'},
    {'category': '개발',
     'name': '권한·감사 모듈',
     'detail': '감사 로그\n로그인·이벤트 이력\n시스템 로그\nWCAG 검증',
     'mm': 0.4, 'unit_price': 3800000, 'amount': 1520000,
     'note': '기본'},
    {'category': '개발',
     'name': '공통 컴포넌트',
     'detail': '홈·통계·5개 지역 지도\n이메일 발송 큐\n카카오맵·SMTP 연동',
     'mm': 1.0, 'unit_price': 3800000, 'amount': 3800000,
     'note': '고급'},
    {'category': '시범운영',
     'name': '시범운영 (포괄 정액)',
     'detail': '피드백·핫픽스\n운영 매뉴얼 작성\n사용자 교육 1회 (1.4개월)',
     'mm': 1.4, 'unit_price': '포괄 정액', 'amount': 4632727,
     'note': '교육 강화 옵션'},
]

booting_maintenance = [
    {'category': '유지보수 1년차',
     'name': '무상 (구축비 포함)',
     'detail': '버그 수정·서버 장애 대응\n경미한 UI 수정',
     'mm': 0.0, 'unit_price': 0, 'amount': 0,
     'note': '2026.12~2027.12'},
    {'category': '유지보수 2년차',
     'name': 'Basic — 월 500,000원 × 12개월',
     'detail': '장애 대응·정기 점검\n소규모 요청 처리',
     'mm': 1.2, 'unit_price': 5000000, 'amount': 6000000,
     'note': '2027.12~2028.12'},
    {'category': '유지보수 3년차',
     'name': 'Basic — 월 500,000원 × 12개월',
     'detail': '장애 대응·정기 점검\n소규모 요청 처리',
     'mm': 1.2, 'unit_price': 5000000, 'amount': 6000000,
     'note': '2028.12~2029.12'},
]


# ────────────────────────────────────────────
# 앤오픈 (6,700만 원, VAT 포함)
# ────────────────────────────────────────────
onopen_company = {
    'name': '주식회사 앤오픈',
    'biz_no': '567-86-98765',
    'ceo': '이 성 준',
    'address': '서울특별시 마포구 양화로 45, 7층 707호',
    'business': '정보통신업',
    'type': '응용소프트웨어 개발 및 공급',
    'phone': '02-987-6543',
    'fax': '02-987-6544',
    'contact': '박팀장 (010-9876-5432)',
}

onopen_modules = [
    {'category': '분석·설계',
     'name': '사전 준비 및 분석',
     'detail': '요구사항 분석\nUI 및 메뉴 설계\nDB 설계\n소개 페이지\nUX 워크숍·접근성 진단',
     'mm': 1.5, 'unit_price': 4500000, 'amount': 6750000,
     'note': 'PM·아키텍트·UI/UX 협업 (고급)\n+ 접근성 컨설팅'},
    {'category': '개발',
     'name': '회원·인증 모듈 + 비밀번호 보안 강화',
     'detail': '회원가입·로그인\n비밀번호·프로필\n세션 모니터링\n실패 잠금 강화',
     'mm': 0.7, 'unit_price': 3800000, 'amount': 2660000,
     'note': '보안 강화'},
    {'category': '개발',
     'name': '지역문제 제보·게이트키핑 + SLA 임박 표시',
     'detail': '제보·공감투표·댓글\n6단계 워크플로우\nSLA 자동 강조\n이메일 알림',
     'mm': 2.0, 'unit_price': 3800000, 'amount': 7600000,
     'note': 'SLA 모니터링 추가'},
    {'category': '개발',
     'name': '리빙랩 + 게시판 + 산출물 공개/비공개 토글',
     'detail': '등록·활동 타임라인\n산출물 DB·공개 토글\n성공사례·게시판',
     'mm': 1.6, 'unit_price': 3800000, 'amount': 6080000,
     'note': '공개 정책 추가'},
    {'category': '개발',
     'name': '멘토단·학생팀 매칭 모듈',
     'detail': '멘토 자격 부여\n학생팀 구성·매칭\n이메일 통보',
     'mm': 0.8, 'unit_price': 3800000, 'amount': 3040000,
     'note': '풀스택 중급'},
    {'category': '개발',
     'name': '협력 네트워크 모듈',
     'detail': '협력기관·MOU\n프로그램·커뮤니티',
     'mm': 0.5, 'unit_price': 3800000, 'amount': 1900000,
     'note': '풀스택 중급'},
    {'category': '개발',
     'name': '성과자료 + 다운로드 통계 대시보드',
     'detail': '성과지표·실적\n자동 집계\n다운로드 통계 시각화',
     'mm': 0.7, 'unit_price': 3800000, 'amount': 2660000,
     'note': '대시보드 강화'},
    {'category': '개발',
     'name': '콘텐츠 관리 모듈',
     'detail': '공지·이벤트 WYSIWYG\n자료실 카테고리·카운트\n배너·약관 버전',
     'mm': 1.4, 'unit_price': 3800000, 'amount': 5320000,
     'note': '풀스택 중급'},
    {'category': '개발',
     'name': '권한·감사 + WCAG 2.1 AA 풀 검증',
     'detail': '감사 로그\n로그인 이력\nWCAG (Lighthouse + axe DevTools)',
     'mm': 0.5, 'unit_price': 3800000, 'amount': 1900000,
     'note': '접근성 풀 검증'},
    {'category': '개발',
     'name': '공통 컴포넌트',
     'detail': '홈·통계\n5개 지역 지도\n이메일 큐·카카오맵·SMTP',
     'mm': 1.0, 'unit_price': 3800000, 'amount': 3800000,
     'note': '고급'},
    {'category': '시범운영',
     'name': '시범운영 2개월·교육·운영 매뉴얼·핫픽스 (포괄 정액)',
     'detail': 'UAT 2개월 강화\n사용자 교육 2회\n운영 매뉴얼\n핫픽스',
     'mm': 1.5, 'unit_price': '포괄 정액', 'amount': 7199091,
     'note': 'UAT 강화 옵션'},
]

onopen_maintenance = [
    {'category': '유지보수 1년차',
     'name': '무상 (구축비 포함)',
     'detail': '버그 수정·서버 장애 대응\n경미한 UI 수정',
     'mm': 0.0, 'unit_price': 0, 'amount': 0,
     'note': '2026.12~2027.12'},
    {'category': '유지보수 2년차',
     'name': 'Standard — 월 500,000원 × 12개월',
     'detail': '장애 대응·정기 점검\n소규모 요청 처리\n분기별 보안 점검',
     'mm': 1.2, 'unit_price': 5000000, 'amount': 6000000,
     'note': '2027.12~2028.12'},
    {'category': '유지보수 3년차',
     'name': 'Standard — 월 500,000원 × 12개월',
     'detail': '장애 대응·정기 점검\n소규모 요청 처리\n분기별 보안 점검',
     'mm': 1.2, 'unit_price': 5000000, 'amount': 6000000,
     'note': '2028.12~2029.12'},
]


# ────────────────────────────────────────────
# Generate
# ────────────────────────────────────────────
if __name__ == '__main__':
    create_quote(
        output_path='/Users/sangincha/dev/deep-sos/USCP_견적서_부팅_20260513.xlsx',
        company=booting_company,
        quote_date='2026년 5월 13일',
        kor_amount='육천삼백만원정',
        modules=booting_modules,
        maintenance=booting_maintenance,
        supply_total=57272727,
        vat=5727273,
        grand_total=63000000,
    )

    create_quote(
        output_path='/Users/sangincha/dev/deep-sos/USCP_견적서_앤오픈_20260514.xlsx',
        company=onopen_company,
        quote_date='2026년 5월 14일',
        kor_amount='육천칠백만원정',
        modules=onopen_modules,
        maintenance=onopen_maintenance,
        supply_total=60909091,
        vat=6090909,
        grand_total=67000000,
    )

    # 검증
    booting_sum = sum(m['amount'] for m in booting_modules) + \
                   sum(m['amount'] for m in booting_maintenance)
    onopen_sum = sum(m['amount'] for m in onopen_modules) + \
                  sum(m['amount'] for m in onopen_maintenance)

    print(f"\n=== 검증 ===")
    print(f"부팅 모듈+유지보수 합계: ₩{booting_sum:,} (공급가 ₩57,272,727 정확={booting_sum==57272727})")
    print(f"앤오픈 모듈+유지보수 합계: ₩{onopen_sum:,} (공급가 ₩60,909,091 정확={onopen_sum==60909091})")
